"""Parser de listas Excel de proveedores (sin I/O de DB)."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.excepciones import ReglaDeNegocioViolada

CAMPOS_MAPEO = {
    "codigo_producto",
    "descripcion",
    "precio_costo",
    "precio_lista",
    "marca",
    "rubro",
    "ignorar",
}

# Etiquetas de UI → campo canónico
ETIQUETA_A_CAMPO: dict[str, str] = {
    "código de producto": "codigo_producto",
    "codigo de producto": "codigo_producto",
    "descripción": "descripcion",
    "descripcion": "descripcion",
    "precio de costo": "precio_costo",
    "precio de lista": "precio_lista",
    "marca": "marca",
    "rubro": "rubro",
    "unidad de medida": "ignorar",
    "alícuota iva": "ignorar",
    "alicuota iva": "ignorar",
    "(ignorar columna)": "ignorar",
    "ignorar columna": "ignorar",
    "ignorar": "ignorar",
}


def normalizar_campo(valor: str) -> str:
    """Convierte etiqueta de UI o código canónico a campo interno."""
    limpio = (valor or "").strip()
    if limpio in CAMPOS_MAPEO:
        return limpio
    clave = limpio.lower()
    if clave in ETIQUETA_A_CAMPO:
        return ETIQUETA_A_CAMPO[clave]
    raise ReglaDeNegocioViolada(f"Campo de mapeo desconocido: {valor}")


def columna_a_indice(columna: str) -> int:
    """'A' / 'Columna A' / 'a' → índice 0-based."""
    texto = (columna or "").strip().upper().replace("COLUMNA", "").strip()
    if not texto or not texto.isalpha():
        raise ReglaDeNegocioViolada(f"Columna Excel inválida: {columna}")
    indice = 0
    for char in texto:
        indice = indice * 26 + (ord(char) - ord("A") + 1)
    return indice - 1


def parsear_numero(valor: Any) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    if not texto:
        return None
    # Formato AR: 61.900,00 → 61900.00
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError as exc:
        raise ReglaDeNegocioViolada(f"Precio inválido: {valor}") from exc


@dataclass(frozen=True)
class FilaListaProveedor:
    sku: str
    nombre: str
    costo: float
    precio_lista: float | None
    marca: str
    rubro: str
    fila_excel: int


@dataclass(frozen=True)
class ResultadoParseoExcel:
    filas: list[FilaListaProveedor]
    preview_cols: list[str]
    preview_rows: list[list[str]]
    omitidas: list[str]


def parsear_lista_excel(
    contenido: bytes,
    mapeo: list[dict[str, str]],
    fila_inicio: int,
    *,
    max_preview: int = 20,
) -> ResultadoParseoExcel:
    """Lee el Excel y materializa filas tipadas según el mapeo."""
    if fila_inicio < 1:
        raise ReglaDeNegocioViolada("La fila de inicio debe ser ≥ 1")
    if not mapeo:
        raise ReglaDeNegocioViolada("El mapeo de columnas es obligatorio")

    campos: dict[str, int] = {}
    for item in mapeo:
        campo = normalizar_campo(item.get("campo", ""))
        if campo == "ignorar":
            continue
        col = columna_a_indice(item.get("columna", ""))
        if campo in campos:
            raise ReglaDeNegocioViolada(f"El campo '{campo}' está mapeado más de una vez")
        campos[campo] = col

    if "codigo_producto" not in campos:
        raise ReglaDeNegocioViolada("El mapeo debe incluir Código de producto")
    if "precio_costo" not in campos:
        raise ReglaDeNegocioViolada("El mapeo debe incluir Precio de costo")

    try:
        libro = load_workbook(filename=BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl lanza varios tipos
        raise ReglaDeNegocioViolada("No se pudo leer el archivo Excel") from exc

    hoja = libro.active
    if hoja is None:
        raise ReglaDeNegocioViolada("El Excel no tiene hojas")

    indices_usados = sorted(campos.values())
    labels = {v: k for k, v in campos.items()}
    preview_cols = [f"{_indice_a_letra(i)} → {labels[i]}" for i in indices_usados]

    filas: list[FilaListaProveedor] = []
    preview_rows: list[list[str]] = []
    omitidas: list[str] = []

    for nro, row in enumerate(hoja.iter_rows(min_row=fila_inicio, values_only=True), start=fila_inicio):
        valores = list(row) if row else []

        def celda(idx: int) -> Any:
            return valores[idx] if idx < len(valores) else None

        sku_raw = celda(campos["codigo_producto"])
        sku = str(sku_raw).strip() if sku_raw is not None else ""
        if not sku:
            continue

        try:
            costo = parsear_numero(celda(campos["precio_costo"]))
        except ReglaDeNegocioViolada:
            omitidas.append(f"Fila {nro}: precio de costo inválido ({sku})")
            continue
        if costo is None or costo < 0:
            omitidas.append(f"Fila {nro}: sin precio de costo ({sku})")
            continue

        nombre = ""
        if "descripcion" in campos:
            desc = celda(campos["descripcion"])
            nombre = str(desc).strip() if desc is not None else ""

        precio_lista = None
        if "precio_lista" in campos:
            try:
                precio_lista = parsear_numero(celda(campos["precio_lista"]))
            except ReglaDeNegocioViolada:
                omitidas.append(f"Fila {nro}: precio de lista inválido ({sku})")
                continue

        marca = ""
        if "marca" in campos:
            m = celda(campos["marca"])
            marca = str(m).strip() if m is not None else ""

        rubro = ""
        if "rubro" in campos:
            r = celda(campos["rubro"])
            rubro = str(r).strip() if r is not None else ""

        filas.append(
            FilaListaProveedor(
                sku=sku[:40],
                nombre=nombre[:120],
                costo=round(costo, 2),
                precio_lista=round(precio_lista, 2) if precio_lista is not None else None,
                marca=marca[:80],
                rubro=rubro[:80],
                fila_excel=nro,
            )
        )
        if len(preview_rows) < max_preview:
            preview_rows.append(
                [
                    _fmt_celda(celda(i))
                    for i in indices_usados
                ]
            )

    libro.close()
    return ResultadoParseoExcel(
        filas=filas,
        preview_cols=preview_cols,
        preview_rows=preview_rows,
        omitidas=omitidas,
    )


def _indice_a_letra(indice: int) -> str:
    n = indice + 1
    letras = ""
    while n:
        n, r = divmod(n - 1, 26)
        letras = chr(65 + r) + letras
    return letras


def _fmt_celda(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float):
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return str(valor)
